"""Quick Analysis API — on-demand Tender Intelligence Report for any opportunity.

Triggered manually from the dashboard. Produces a v2.0 structured report
using opportunity metadata + description text + document content.
Stores the full report JSON in tender_intelligence.intelligence_summary.

Key behaviour:
  - On-demand document extraction: un-extracted documents are extracted
    inline BEFORE calling OpenAI so the AI always sees all available content.
  - Web-page content extraction: link-type documents have their page text
    fetched and stored so the AI can read external resources too.
  - Cost control: AI analysis is ONLY user-triggered, never automatic.
    Budget limits are enforced per-day and per-month.
"""

from __future__ import annotations

import io
import json
import re
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

import requests as http_requests
from fastapi import APIRouter, Depends, HTTPException, status
from pydantic import BaseModel
from sqlalchemy import text

from src.api.auth import verify_api_key
from src.core.config import settings
from src.core.database import get_db_session
from src.core.logging import get_logger
from src.intelligence.analyzer import TenderAnalyzer

logger = get_logger(__name__)
router = APIRouter(prefix="/api/analysis", tags=["analysis"])

_COST_PER_1K_INPUT: dict[str, float] = {
    "gpt-4o-mini": 0.00015,
    "gpt-4o": 0.0025,
}
_COST_PER_1K_OUTPUT: dict[str, float] = {
    "gpt-4o-mini": 0.0006,
    "gpt-4o": 0.01,
}


def _estimate_cost(model: str, prompt_tokens: int, completion_tokens: int) -> float:
    input_cost = (prompt_tokens / 1000) * _COST_PER_1K_INPUT.get(model, 0.003)
    output_cost = (completion_tokens / 1000) * _COST_PER_1K_OUTPUT.get(model, 0.006)
    return round(input_cost + output_cost, 6)


def _check_budget(session: Any, mode: str) -> tuple[bool, str]:
    """Check daily and monthly AI budget limits. Returns (ok, message)."""
    try:
        today_row = session.execute(
            text("""
                SELECT COALESCE(SUM(estimated_cost_usd), 0) as total
                FROM ai_usage_log
                WHERE created_at >= CURRENT_DATE
            """),
        ).fetchone()
        daily_spent = float(today_row.total) if today_row else 0.0

        month_row = session.execute(
            text("""
                SELECT COALESCE(SUM(estimated_cost_usd), 0) as total
                FROM ai_usage_log
                WHERE created_at >= date_trunc('month', CURRENT_DATE)
            """),
        ).fetchone()
        monthly_spent = float(month_row.total) if month_row else 0.0

        if daily_spent >= settings.AI_DAILY_BUDGET_USD:
            return False, f"Daily AI budget reached (${daily_spent:.2f} / ${settings.AI_DAILY_BUDGET_USD:.2f}). Try again tomorrow."
        if monthly_spent >= settings.AI_MONTHLY_BUDGET_USD:
            return False, f"Monthly AI budget reached (${monthly_spent:.2f} / ${settings.AI_MONTHLY_BUDGET_USD:.2f}). Contact admin."

        return True, ""
    except Exception as exc:
        logger.warning("Budget check failed (allowing analysis): %s", exc)
        return True, ""


def _record_usage(
    session: Any,
    opportunity_id: str,
    model: str,
    mode: str,
    prompt_tokens: int,
    completion_tokens: int,
    estimated_cost: float,
    ts: datetime,
) -> None:
    """Log AI usage for budget tracking."""
    try:
        session.execute(
            text("""
                INSERT INTO ai_usage_log (
                    opportunity_id, model, analysis_mode,
                    prompt_tokens, completion_tokens, total_tokens,
                    estimated_cost_usd, created_at
                ) VALUES (
                    :opp_id, :model, :mode,
                    :prompt, :completion, :total,
                    :cost, :ts
                )
            """),
            {
                "opp_id": opportunity_id,
                "model": model,
                "mode": mode,
                "prompt": prompt_tokens,
                "completion": completion_tokens,
                "total": prompt_tokens + completion_tokens,
                "cost": estimated_cost,
                "ts": ts,
            },
        )
    except Exception as exc:
        logger.warning("Failed to record AI usage: %s", exc)


_EXTRACT_TIMEOUT = 45
_EXTRACT_MAX_SIZE = 20 * 1024 * 1024


def _extract_text_from_bytes(content: bytes, file_type: str) -> str:
    """Extract text from downloaded document bytes. Mirrors extract_documents.py extractors."""
    ft = file_type.lower().strip(".")
    if ft == "pdf":
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(content))
        parts = [p.extract_text() or "" for p in reader.pages]
        return "\n\n".join(p for p in parts if p)
    if ft in ("docx", "doc"):
        import tempfile
        from docx import Document
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
            tmp.write(content)
            tmp_path = Path(tmp.name)
        try:
            doc = Document(str(tmp_path))
            return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
        finally:
            tmp_path.unlink(missing_ok=True)
    if ft in ("xlsx", "xls"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet in wb.worksheets:
            parts.append(f"--- Sheet: {sheet.title} ---")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    parts.append("\t".join(cells))
        wb.close()
        return "\n".join(parts)
    if ft == "csv":
        import csv as csv_mod
        text_str = content.decode("utf-8", errors="replace")
        reader = csv_mod.reader(io.StringIO(text_str))
        return "\n".join("\t".join(row) for row in reader)
    if ft == "txt":
        return content.decode("utf-8", errors="replace")
    return ""


def _extract_web_page(url: str) -> str:
    """Fetch a web page and extract its meaningful text content."""
    try:
        from bs4 import BeautifulSoup
        resp = http_requests.get(url, timeout=_EXTRACT_TIMEOUT, headers={
            "User-Agent": "Mozilla/5.0 (compatible; BidToGo/1.0)"
        })
        resp.raise_for_status()
        ct = resp.headers.get("content-type", "")
        if "pdf" in ct:
            return _extract_text_from_bytes(resp.content, "pdf")
        if "html" not in ct and "text" not in ct:
            return ""
        soup = BeautifulSoup(resp.text, "lxml")
        for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
            tag.decompose()
        text_str = soup.get_text(separator="\n", strip=True)
        lines = [ln.strip() for ln in text_str.splitlines() if ln.strip()]
        return "\n".join(lines)[:50000]
    except Exception as exc:
        logger.warning("Failed to extract web page %s: %s", url[:80], exc)
        return ""


def _ensure_documents_extracted(session: Any, opportunity_id: str) -> dict:
    """Extract text from all un-extracted documents for an opportunity, inline.

    Returns summary dict with counts.
    """
    docs = session.execute(
        text("""
            SELECT id, title, url, file_type
            FROM opportunity_documents
            WHERE opportunity_id = :opp_id
              AND (text_extracted = false OR extracted_text IS NULL)
              AND url IS NOT NULL AND url != ''
            ORDER BY created_at
        """),
        {"opp_id": opportunity_id},
    ).fetchall()

    if not docs:
        return {"extracted": 0, "skipped": 0, "failed": 0, "web_pages": 0}

    extracted = skipped = failed = web_pages = 0

    for doc in docs:
        file_type = (doc.file_type or "").lower().strip(".")
        is_link = file_type in ("link", "html", "htm", "")

        try:
            if is_link and doc.url:
                page_text = _extract_web_page(doc.url)
                if page_text and len(page_text.strip()) > 20:
                    page_text = page_text[:200000]
                    session.execute(
                        text("""
                            UPDATE opportunity_documents SET
                                extracted_text = :txt,
                                text_extracted = true,
                                page_count = 1
                            WHERE id = :id
                        """),
                        {"id": doc.id, "txt": page_text},
                    )
                    session.commit()
                    web_pages += 1
                    logger.info("Extracted %d chars from web page '%s'", len(page_text), (doc.title or doc.url)[:60])
                else:
                    session.execute(
                        text("UPDATE opportunity_documents SET text_extracted = true WHERE id = :id"),
                        {"id": doc.id},
                    )
                    session.commit()
                    skipped += 1
                continue

            supported = {"pdf", "docx", "doc", "txt", "xlsx", "xls", "csv"}
            if file_type not in supported:
                skipped += 1
                continue

            resp = http_requests.get(doc.url, timeout=_EXTRACT_TIMEOUT, stream=True)
            resp.raise_for_status()

            content_length = int(resp.headers.get("content-length", 0))
            if content_length > _EXTRACT_MAX_SIZE:
                skipped += 1
                continue

            content = resp.content
            if len(content) > _EXTRACT_MAX_SIZE:
                skipped += 1
                continue

            text_content = _extract_text_from_bytes(content, file_type)

            if not text_content or len(text_content.strip()) < 10:
                session.execute(
                    text("UPDATE opportunity_documents SET text_extracted = true WHERE id = :id"),
                    {"id": doc.id},
                )
                session.commit()
                skipped += 1
                continue

            text_content = text_content[:200000]
            session.execute(
                text("""
                    UPDATE opportunity_documents SET
                        extracted_text = :txt,
                        text_extracted = true,
                        page_count = :pc
                    WHERE id = :id
                """),
                {
                    "id": doc.id,
                    "txt": text_content,
                    "pc": text_content.count("\n\n") + 1,
                },
            )
            session.commit()
            extracted += 1
            logger.info("On-demand extracted %d chars from '%s' (%s)",
                        len(text_content), doc.title or doc.id, file_type)

        except Exception as exc:
            logger.warning("On-demand extraction failed for doc %s: %s", doc.id, exc)
            failed += 1

    result = {"extracted": extracted, "skipped": skipped, "failed": failed, "web_pages": web_pages}
    logger.info("On-demand extraction complete for %s: %s", opportunity_id, result)
    return result


class _Enc(json.JSONEncoder):
    def default(self, o: object) -> object:
        if isinstance(o, (datetime, date)):
            return o.isoformat()
        if isinstance(o, Decimal):
            return float(o)
        return super().default(o)


class AnalyzeRequest(BaseModel):
    opportunity_id: str
    mode: str = "quick"


class AnalyzeResponse(BaseModel):
    status: str
    opportunity_id: str
    overall_score: int | None = None
    recommendation: str | None = None
    confidence: str | None = None
    analysis_model: str | None = None
    message: str | None = None
    documents_analyzed: int | None = None
    total_documents: int | None = None


@router.post("/run", dependencies=[Depends(verify_api_key)])
async def run_quick_analysis(req: AnalyzeRequest) -> AnalyzeResponse:
    """Run on-demand Quick Analysis and produce a Tender Intelligence Report."""
    session = get_db_session()
    try:
        opp = session.execute(
            text("""
                SELECT o.id, o.title, o.description_summary, o.description_full,
                       o.country, o.region, o.city, o.closing_date, o.source_url,
                       o.relevance_score, o.relevance_bucket, o.keywords_matched,
                       o.industry_tags, o.category, o.project_type,
                       o.solicitation_number, o.contact_name, o.contact_email,
                       o.raw_data,
                       s.name as source_name,
                       org.name as organization_name
                FROM opportunities o
                LEFT JOIN sources s ON o.source_id = s.id
                LEFT JOIN organizations org ON o.organization_id = org.id
                WHERE o.id = :id
            """),
            {"id": req.opportunity_id},
        ).fetchone()

        if not opp:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Opportunity not found")

        existing = session.execute(
            text("SELECT id, analyzed_at FROM tender_intelligence WHERE opportunity_id = :id"),
            {"id": req.opportunity_id},
        ).fetchone()

        if existing:
            logger.info("Re-analyzing opportunity %s (previous at %s)", req.opportunity_id, existing.analyzed_at)

        description = opp.description_full or opp.description_summary or ""
        location_parts = [p for p in [opp.city, opp.region, opp.country] if p]
        location = ", ".join(location_parts) if location_parts else None

        raw = opp.raw_data if opp.raw_data else {}
        if isinstance(raw, str):
            try:
                raw = json.loads(raw)
            except Exception:
                raw = {}

        extraction_result = _ensure_documents_extracted(session, req.opportunity_id)
        logger.info("Document extraction summary: %s", extraction_result)

        doc_rows = session.execute(
            text("""
                SELECT id, title, url, file_type, file_size_bytes, extracted_text
                FROM opportunity_documents
                WHERE opportunity_id = :id AND text_extracted = true
                  AND extracted_text IS NOT NULL AND LENGTH(extracted_text) > 10
                ORDER BY
                    CASE WHEN LOWER(file_type) IN ('pdf','docx','doc','xlsx','xls') THEN 0 ELSE 1 END,
                    file_size_bytes DESC NULLS LAST
                LIMIT 15
            """),
            {"id": req.opportunity_id},
        ).fetchall()
        document_texts: dict[str, str] | None = None
        docs_used: list[dict] = []
        if doc_rows:
            document_texts = {}
            for i, row in enumerate(doc_rows):
                key = row.title or f"document_{i}"
                document_texts[key] = row.extracted_text
                docs_used.append({
                    "id": str(row.id),
                    "title": row.title,
                    "file_type": row.file_type,
                    "chars": len(row.extracted_text or ""),
                })
            logger.info("Including %d document(s) with extracted text for analysis", len(document_texts))

        total_docs_count = session.execute(
            text("SELECT COUNT(*) as cnt FROM opportunity_documents WHERE opportunity_id = :id"),
            {"id": req.opportunity_id},
        ).fetchone()
        total_docs = total_docs_count.cnt if total_docs_count else 0

        budget_ok, budget_msg = _check_budget(session, req.mode)
        if not budget_ok:
            logger.warning("AI budget exceeded: %s", budget_msg)
            return AnalyzeResponse(
                status="budget_exceeded",
                opportunity_id=req.opportunity_id,
                message=budget_msg,
            )

        use_model = "gpt-4o" if req.mode == "deep" else "gpt-4o-mini"
        max_tok = 8000 if req.mode == "deep" else 4000
        analyzer = TenderAnalyzer(model=use_model, max_tokens=max_tok)
        result = analyzer.analyze(
            title=opp.title,
            organization=opp.organization_name,
            location=location,
            closing_date=str(opp.closing_date) if opp.closing_date else None,
            source=opp.source_name or "Unknown",
            description=description,
            document_texts=document_texts,
            country=opp.country,
            response_deadline=raw.get("response_deadline"),
            naics=raw.get("naics_code") or opp.category,
            category=opp.category,
            set_aside=raw.get("set_aside"),
            solicitation_number=opp.solicitation_number,
        )

        fallback_used = result.get("fallback_used", False)
        verdict = result.get("verdict", {})
        scores = result.get("feasibility_scores", {})
        overall = scores.get("overall_score")
        recommendation = verdict.get("recommendation", "review_carefully")
        confidence = verdict.get("confidence", "low")
        analysis_model = result.get("analysis_model", "gpt-4o-mini")
        now = datetime.now(timezone.utc)

        prompt_tok = result.get("_prompt_tokens", 0)
        completion_tok = result.get("_completion_tokens", 0)
        est_cost = _estimate_cost(use_model, prompt_tok, completion_tok) if not fallback_used else 0.0

        if not fallback_used and est_cost > 0:
            _record_usage(session, req.opportunity_id, use_model, req.mode, prompt_tok, completion_tok, est_cost, now)

        if fallback_used:
            logger.warning(
                "Analysis used FALLBACK for opp=%s — OpenAI did not produce the report",
                req.opportunity_id,
            )

        result.pop("_prompt_tokens", None)
        result.pop("_completion_tokens", None)

        result["_analysis_metadata"] = {
            "documents_used": docs_used,
            "documents_used_count": len(docs_used),
            "total_documents": total_docs,
            "total_doc_chars": sum(d["chars"] for d in docs_used),
            "extraction_summary": extraction_result,
        }

        biz_fit = result.get("business_fit", {})
        scope = result.get("scope_breakdown", {})
        tech = result.get("technical_requirements", {})
        quals = result.get("compliance_risks", {})
        timeline = result.get("timeline_milestones", {})
        risks_list = [rf.get("requirement", "") for rf in result.get("compliance_risks", {}).get("red_flags", [])]
        china = result.get("supply_chain_feasibility", {})

        params = {
            "opp_id": req.opportunity_id,
            "overview": result.get("project_summary", {}).get("overview", ""),
            "scope": json.dumps(scope, cls=_Enc),
            "scope_type": scope.get("scope_type", "unclear"),
            "tech_reqs": json.dumps(tech, cls=_Enc),
            "qual_reqs": json.dumps(quals, cls=_Enc),
            "dates": json.dumps(timeline, cls=_Enc),
            "risks": json.dumps(risks_list, cls=_Enc),
            "feas_score": overall,
            "recommendation": recommendation,
            "biz_fit": biz_fit.get("fit_explanation", "")[:500],
            "china": json.dumps(china, cls=_Enc),
            "summary": json.dumps(result, cls=_Enc),
            "model": analysis_model,
            "mode": req.mode,
            "status": "completed" if not fallback_used else "fallback_only",
            "now": now,
        }

        if existing:
            session.execute(
                text("""
                    UPDATE tender_intelligence SET
                        project_overview = :overview,
                        scope_of_work = :scope,
                        scope_type = :scope_type,
                        technical_requirements = :tech_reqs,
                        qualification_reqs = :qual_reqs,
                        critical_dates = :dates,
                        risk_factors = :risks,
                        feasibility_score = :feas_score,
                        recommendation_status = :recommendation,
                        business_fit_explanation = :biz_fit,
                        china_source_analysis = :china,
                        intelligence_summary = :summary,
                        analysis_model = :model,
                        analysis_mode = :mode,
                        analysis_status = :status,
                        analyzed_at = :now,
                        updated_at = :now
                    WHERE opportunity_id = :opp_id
                """),
                params,
            )
        else:
            session.execute(
                text("""
                    INSERT INTO tender_intelligence (
                        opportunity_id, project_overview, scope_of_work, scope_type,
                        technical_requirements, qualification_reqs, critical_dates,
                        risk_factors, feasibility_score, recommendation_status,
                        business_fit_explanation, china_source_analysis,
                        intelligence_summary, analysis_model, analysis_mode,
                        analysis_status, analyzed_at, updated_at
                    ) VALUES (
                        :opp_id, :overview, :scope, :scope_type,
                        :tech_reqs, :qual_reqs, :dates,
                        :risks, :feas_score, :recommendation,
                        :biz_fit, :china,
                        :summary, :model, :mode,
                        :status, :now, :now
                    )
                """),
                params,
            )

        session.execute(
            text("UPDATE opportunities SET business_fit_explanation = :biz, updated_at = :now WHERE id = :id"),
            {"id": req.opportunity_id, "biz": biz_fit.get("fit_explanation", "")[:500], "now": now},
        )
        session.commit()

        logger.info(
            "Tender Intelligence Report complete: opp=%s score=%s rec=%s conf=%s model=%s",
            req.opportunity_id, overall, recommendation, confidence, analysis_model,
        )

        return AnalyzeResponse(
            status="completed",
            opportunity_id=req.opportunity_id,
            overall_score=overall,
            recommendation=recommendation,
            confidence=confidence,
            analysis_model=analysis_model,
            documents_analyzed=len(docs_used),
            total_documents=total_docs,
        )

    except HTTPException:
        raise
    except Exception as exc:
        session.rollback()
        logger.exception("Quick analysis failed for %s", req.opportunity_id)
        return AnalyzeResponse(
            status="failed",
            opportunity_id=req.opportunity_id,
            message=str(exc),
        )
    finally:
        session.close()


@router.get("/status/{opportunity_id}", dependencies=[Depends(verify_api_key)])
async def analysis_status(opportunity_id: str) -> dict:
    """Check if an analysis exists for an opportunity."""
    session = get_db_session()
    try:
        row = session.execute(
            text("""
                SELECT id, feasibility_score, recommendation_status,
                       analysis_model, analyzed_at
                FROM tender_intelligence
                WHERE opportunity_id = :id
            """),
            {"id": opportunity_id},
        ).fetchone()
        if row:
            return {
                "exists": True,
                "intel_id": str(row.id),
                "feasibility_score": row.feasibility_score,
                "recommendation": row.recommendation_status,
                "model": row.analysis_model,
                "analyzed_at": row.analyzed_at.isoformat() if row.analyzed_at else None,
            }
        return {"exists": False}
    finally:
        session.close()


# ──────────────────────────────────────────────────────────────
# Upload & Analyze — manual PDF upload for deep analysis
# ──────────────────────────────────────────────────────────────

from fastapi import File, Form, UploadFile

_UPLOAD_MAX_FILES = 5
_UPLOAD_MAX_SIZE = 25 * 1024 * 1024  # 25 MB per file


@router.post("/upload-and-analyze", dependencies=[Depends(verify_api_key)])
async def upload_and_analyze(
    files: list[UploadFile] = File(...),
    opportunity_id: str | None = Form(None),
    title: str | None = Form(None),
) -> dict:
    """Upload 1-5 PDF/DOCX files and run deep AI analysis.

    If opportunity_id is provided, enriches the analysis with opportunity metadata.
    Otherwise runs a standalone analysis using only the uploaded documents.
    """
    if len(files) > _UPLOAD_MAX_FILES:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"最多上传 {_UPLOAD_MAX_FILES} 个文件")

    document_texts: dict[str, str] = {}
    for f in files:
        if not f.filename:
            continue
        content = await f.read()
        if len(content) > _UPLOAD_MAX_SIZE:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, f"文件 {f.filename} 超过 25MB 限制")
        ext = (f.filename.rsplit(".", 1)[-1] if "." in f.filename else "").lower()
        supported = {"pdf", "docx", "doc", "txt", "xlsx", "xls", "csv"}
        if ext not in supported:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"不支持的文件类型: {f.filename}（支持: {', '.join(supported)}）",
            )
        try:
            extracted = _extract_text_from_bytes(content, ext)
            if extracted and len(extracted.strip()) > 10:
                document_texts[f.filename] = extracted
        except Exception as exc:
            logger.warning("Failed to extract text from uploaded %s: %s", f.filename, exc)

    if not document_texts:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "未能从上传的文件中提取到有效文本")

    session = get_db_session()
    try:
        budget_ok, budget_msg = _check_budget(session, "upload_deep")
        if not budget_ok:
            return {"status": "budget_exceeded", "message": budget_msg}

        opp_title = title or "上传文档分析"
        organization = None
        location = None
        closing_date = None
        source_name = "手动上传"
        description = ""
        country = None
        naics = None
        category = None
        set_aside = None
        solicitation_number = None
        response_deadline = None
        linked_opp_id = opportunity_id

        if opportunity_id:
            opp = session.execute(
                text("""
                    SELECT o.id, o.title, o.description_summary, o.description_full,
                           o.country, o.region, o.city, o.closing_date, o.category,
                           o.solicitation_number, o.raw_data,
                           s.name as source_name,
                           org.name as organization_name
                    FROM opportunities o
                    LEFT JOIN sources s ON o.source_id = s.id
                    LEFT JOIN organizations org ON o.organization_id = org.id
                    WHERE o.id = :id
                """),
                {"id": opportunity_id},
            ).fetchone()

            if opp:
                opp_title = opp.title
                organization = opp.organization_name
                lp = [p for p in [opp.city, opp.region, opp.country] if p]
                location = ", ".join(lp) if lp else None
                closing_date = str(opp.closing_date) if opp.closing_date else None
                source_name = opp.source_name or "Unknown"
                description = opp.description_full or opp.description_summary or ""
                country = opp.country
                category = opp.category
                solicitation_number = opp.solicitation_number
                raw = opp.raw_data if opp.raw_data else {}
                if isinstance(raw, str):
                    try:
                        raw = json.loads(raw)
                    except Exception:
                        raw = {}
                naics = raw.get("naics_code") or opp.category
                set_aside = raw.get("set_aside")
                response_deadline = raw.get("response_deadline")

            existing_docs = session.execute(
                text("""
                    SELECT title, extracted_text FROM opportunity_documents
                    WHERE opportunity_id = :id
                      AND text_extracted = true
                      AND extracted_text IS NOT NULL
                      AND LENGTH(extracted_text) > 10
                    LIMIT 10
                """),
                {"id": opportunity_id},
            ).fetchall()
            for i, doc in enumerate(existing_docs):
                key = doc.title or f"existing_doc_{i}"
                if key not in document_texts:
                    document_texts[key] = doc.extracted_text

        analyzer = TenderAnalyzer(model="gpt-4o", max_tokens=AUTO_ANALYZE_MAX_TOKENS)
        result = analyzer.analyze(
            title=opp_title,
            organization=organization,
            location=location,
            closing_date=closing_date,
            source=source_name,
            description=description,
            document_texts=document_texts,
            country=country,
            response_deadline=response_deadline,
            naics=naics,
            category=category,
            set_aside=set_aside,
            solicitation_number=solicitation_number,
        )

        fallback_used = result.get("fallback_used", False)
        now = datetime.now(timezone.utc)
        prompt_tok = result.pop("_prompt_tokens", 0)
        completion_tok = result.pop("_completion_tokens", 0)

        if not fallback_used:
            est_cost = _estimate_cost("gpt-4o", prompt_tok, completion_tok)
            _record_usage(
                session, linked_opp_id or "upload", "gpt-4o", "upload_deep",
                prompt_tok, completion_tok, est_cost, now,
            )

        if linked_opp_id:
            verdict = result.get("verdict", {})
            scores = result.get("feasibility_scores", {})
            overall = scores.get("overall_score")
            recommendation = verdict.get("recommendation", "review_carefully")
            biz_fit = result.get("business_fit", {})
            scope = result.get("scope_breakdown", {})
            tech = result.get("technical_requirements", {})
            quals = result.get("compliance_risks", {})
            timeline = result.get("timeline_milestones", {})
            risks_list = [rf.get("requirement", "") for rf in quals.get("red_flags", [])]
            china = result.get("supply_chain_feasibility", {})

            params = {
                "opp_id": linked_opp_id,
                "overview": result.get("project_summary", {}).get("overview", ""),
                "scope": json.dumps(scope, cls=_Enc),
                "scope_type": scope.get("scope_type", "unclear"),
                "tech_reqs": json.dumps(tech, cls=_Enc),
                "qual_reqs": json.dumps(quals, cls=_Enc),
                "dates": json.dumps(timeline, cls=_Enc),
                "risks": json.dumps(risks_list, cls=_Enc),
                "feas_score": overall,
                "recommendation": recommendation,
                "biz_fit": biz_fit.get("fit_explanation", "")[:500],
                "china": json.dumps(china, cls=_Enc),
                "summary": json.dumps(result, cls=_Enc),
                "model": "gpt-4o",
                "mode": "upload_deep",
                "status": "completed" if not fallback_used else "fallback_only",
                "now": now,
            }

            existing = session.execute(
                text("SELECT id FROM tender_intelligence WHERE opportunity_id = :id"),
                {"id": linked_opp_id},
            ).fetchone()

            if existing:
                session.execute(text("""
                    UPDATE tender_intelligence SET
                        project_overview = :overview, scope_of_work = :scope,
                        scope_type = :scope_type, technical_requirements = :tech_reqs,
                        qualification_reqs = :qual_reqs, critical_dates = :dates,
                        risk_factors = :risks, feasibility_score = :feas_score,
                        recommendation_status = :recommendation,
                        business_fit_explanation = :biz_fit,
                        china_source_analysis = :china,
                        intelligence_summary = :summary, analysis_model = :model,
                        analysis_mode = :mode, analysis_status = :status,
                        analyzed_at = :now, updated_at = :now
                    WHERE opportunity_id = :opp_id
                """), params)
            else:
                session.execute(text("""
                    INSERT INTO tender_intelligence (
                        opportunity_id, project_overview, scope_of_work, scope_type,
                        technical_requirements, qualification_reqs, critical_dates,
                        risk_factors, feasibility_score, recommendation_status,
                        business_fit_explanation, china_source_analysis,
                        intelligence_summary, analysis_model, analysis_mode,
                        analysis_status, analyzed_at, updated_at
                    ) VALUES (
                        :opp_id, :overview, :scope, :scope_type,
                        :tech_reqs, :qual_reqs, :dates,
                        :risks, :feas_score, :recommendation,
                        :biz_fit, :china,
                        :summary, :model, :mode,
                        :status, :now, :now
                    )
                """), params)

            session.commit()

        return {
            "status": "completed",
            "opportunity_id": linked_opp_id,
            "report": result,
            "documents_analyzed": list(document_texts.keys()),
        }

    except HTTPException:
        raise
    except Exception as exc:
        session.rollback()
        logger.exception("Upload analysis failed")
        return {"status": "error", "message": str(exc)}
    finally:
        session.close()


AUTO_ANALYZE_MAX_TOKENS = 8000
